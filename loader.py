from itertools import islice
from datetime import datetime

from openpyxl import load_workbook
import psycopg2
from psycopg2 import sql
from psycopg2.extensions import AsIs

from settings import get_settings


def parse_file(filename: str = None) -> list:
    if filename:
        book = load_workbook(filename, read_only=True)
        work_sheet = book[book.sheetnames[0]]

        for row in islice(work_sheet.rows, 1, None):
            row_data = []
            cell = None
            try:
                for cell in row:
                    cell = cell
                    value = str(cell.value).replace(',', '.')
                    if cell.column == 5:
                        datetime.fromisoformat(value)
                        row_data.append(value)
                    else:
                        float(value)
                        row_data.append(value)
            except ValueError:
                print(f"Invalid value in cell {cell.coordinate}")
                break
            yield row_data


def upload_data(filename: str = None):
    with psycopg2.connect(dsn=get_settings().db.dsn) as conn:
        with conn.cursor() as cur:

            sql_points_expr = sql.SQL("INSERT INTO {} ({}) VALUES ").format(sql.Identifier('track_point_model'),
                            sql.SQL(', ').join(map(sql.Identifier, ['id', 'point', 'speed', 'gps_time', 'vehicle_id'])))

            sql_vehicles_expr = sql.SQL("INSERT INTO {} ({}) VALUES ").format(
                            sql.Identifier('vehicle_model'), sql.Identifier('vehicle_id'))

            sql_points_expr_values = []
            sql_vehicles_expr_values = []

            for n in parse_file(filename):
                sql_points_expr_values.append(*map(sql.Literal, [AsIs(f"({n[0]}, 'POINT({n[2]} {n[1]})', {n[3]}, '{n[4]}', {n[5]})")]))
                sql_vehicles_expr_values.append(sql.Literal(AsIs(f'({n[5]})')))

            try:
                cur.execute(sql.Composed([sql_vehicles_expr, sql.SQL(', ').join(sql_vehicles_expr_values),
                                          sql.Literal(AsIs('ON CONFLICT DO NOTHING'))]))
                cur.execute(sql.Composed([sql_points_expr, sql.SQL(', ').join(sql_points_expr_values)]))
                conn.commit()
            except psycopg2.Error as e:
                print(e)
                conn.rollback()


if __name__ == '__main__':
    upload_data("")
